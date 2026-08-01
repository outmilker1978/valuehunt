"""
Import weekly retro data from Excel into ValueHunt DB.

Core logic — used by both CLI (scripts/import_retro.py) and API (app.py).
"""
import os, re, json, sys
from pathlib import Path
from typing import Any

import openpyxl

from src.db import get_connection, save_vacancy, _resolve_company_id

# Pipeline order for downgrade protection
PIPELINE = {'new': 0, 'applied': 1, 'invited': 2, 'in_progress': 3, 'offer': 4}
TERMINAL = {'rejected', 'archived', 'closed', 'trash'}

# Each entry: (vh_status_to_set, should_delete, delete_reason)
# vh_status_to_set = None means keep current status when deleting
STATUS_MAP = {
    'отклик': ('applied', False, None),
    'отказ': ('rejected', False, None),
    'архив': (None, True, 'Работодатель закрыл вакансию'),
    'вакансия недоступна к просмотру': (None, True, 'Работодатель закрыл вакансию'),
    'пригласили': ('invited', False, None),
    'удаляю чат': (None, True, 'Работодатель закрыл вакансию'),
}

BASE_DIR = Path(__file__).resolve().parent.parent
EXCEL_PATH = BASE_DIR / 'Аналитика откликов до старта работы.xlsx'


def parse_hh_id(url: str | None) -> str | None:
    if not url:
        return None
    m = re.search(r'vacancy/(\d+)', url)
    return m.group(1) if m else None


def excel_status_to_vh(excel_status: str) -> tuple[str | None, bool, str | None]:
    """Returns (vh_status_to_set, should_delete, delete_reason)."""
    if not excel_status:
        return None, False, None
    s = excel_status.strip().lower()
    for key, (vh, should_del, reason) in STATUS_MAP.items():
        if key in s:
            return vh, should_del, reason
    return None, False, None


def should_update_status(current_status: str | None, incoming_status: str) -> bool:
    if not current_status or not incoming_status:
        return True
    cur = current_status.strip().lower()
    inc = incoming_status.lower()
    if cur in TERMINAL:
        return False
    if inc in TERMINAL:
        return True
    cur_level = PIPELINE.get(cur, -1)
    inc_level = PIPELINE.get(inc, -1)
    return inc_level >= cur_level


def import_retro(dry_run: bool = False, sheet_name: str | None = None) -> dict[str, Any]:
    if not EXCEL_PATH.exists():
        return {'error': f'Excel file not found: {EXCEL_PATH}'}

    wb = openpyxl.load_workbook(str(EXCEL_PATH), data_only=True)
    sheets = wb.sheetnames
    if not sheets:
        return {'error': 'No sheets found in Excel'}

    if sheet_name:
        if sheet_name not in sheets:
            return {'error': f'Sheet "{sheet_name}" not found. Available: {", ".join(sheets)}'}
        target_sheet = sheet_name
    else:
        target_sheet = max(sheets, key=lambda s: s.split('.')[::-1])
    ws = wb[target_sheet]

    conn = get_connection()
    report: dict[str, Any] = {
        'sheet': target_sheet,
        'total_rows': 0,
        'skipped_no_url': 0,
        'created': 0,
        'updated': 0,
        'skipped_downgrade': 0,
        'skipped_terminal': 0,
        'errors': [],
        'contacts_created': 0,
    }

    for row in ws.iter_rows(min_row=16, max_row=ws.max_row, values_only=True):
        company_name, url, excel_status, comment, hr_name = (row + (None,) * 5)[:5]

        if not company_name and not url:
            continue

        company_name = str(company_name).strip() if company_name else ''
        url = str(url).strip() if url else ''
        excel_status = str(excel_status).strip() if excel_status else ''
        comment = str(comment).strip() if comment else ''
        hr_name = str(hr_name).strip() if hr_name else ''

        report['total_rows'] += 1

        hh_id = parse_hh_id(url)
        if not hh_id:
            report['skipped_no_url'] += 1
            report['errors'].append(f'Row ~{report["total_rows"]+15}: no valid URL ({company_name})')
            continue

        vh_status, should_delete, del_reason = excel_status_to_vh(excel_status)

        existing = conn.execute(
            "SELECT id, status, deleted_at, notes, hr_contacts, delete_reason FROM vacancies WHERE hh_id = ?",
            (hh_id,)
        ).fetchone()

        vid = None
        if existing:
            vid = existing['id']
            if existing['deleted_at']:
                report['skipped_terminal'] += 1
                # Still update hr_contacts if new info
                old_hr = existing['hr_contacts'] or ''
                if hr_name and hr_name not in old_hr:
                    new_hr = (old_hr + ', ' + hr_name).strip()
                    if not dry_run:
                        conn.execute("UPDATE vacancies SET hr_contacts = ? WHERE id = ?", (new_hr, vid))
                continue

            cur_status = existing['status']
            old_notes = existing['notes'] or ''
            old_hr = existing['hr_contacts'] or ''

            new_notes = old_notes
            new_hr = old_hr
            if comment and comment not in old_notes:
                new_notes = (old_notes + '\n' + comment).strip()
            if hr_name and hr_name not in old_hr:
                new_hr = (old_hr + ', ' + hr_name).strip()

            if should_delete:
                # Delete action — always apply (terminal outcome)
                if not dry_run:
                    conn.execute(
                        """UPDATE vacancies SET
                            deleted_at = datetime('now'), delete_reason = ?,
                            notes = ?, hr_contacts = ?
                        WHERE id = ?""",
                        (del_reason, new_notes, new_hr, vid)
                    )
                report['updated'] += 1
            elif vh_status and should_update_status(cur_status, vh_status):
                # Status progression
                if not dry_run:
                    conn.execute(
                        """UPDATE vacancies SET
                            status = ?, notes = ?, hr_contacts = ?,
                            responded_at = CASE WHEN ? = 'applied' AND responded_at IS NULL THEN datetime('now') ELSE responded_at END
                        WHERE id = ?""",
                        (vh_status, new_notes, new_hr, vh_status, vid)
                    )
                report['updated'] += 1
            else:
                # No status change — still update notes/hr if changed
                report['skipped_downgrade'] += 1
                if (new_notes != old_notes or new_hr != old_hr) and not dry_run:
                    conn.execute(
                        "UPDATE vacancies SET notes = ?, hr_contacts = ? WHERE id = ?",
                        (new_notes, new_hr, vid)
                    )
        else:
            # New vacancy
            if not dry_run:
                set_status = vh_status or 'applied'
                result = save_vacancy(conn, {
                    'hh_id': hh_id,
                    'title': company_name or '',
                    'company': company_name,
                    'url': url,
                    'status': set_status,
                    'hr_contacts': hr_name or None,
                })
                vid = result['id']
                if set_status == 'applied':
                    conn.execute("UPDATE vacancies SET responded_at = datetime('now') WHERE id = ?", (vid,))
                if comment:
                    conn.execute("UPDATE vacancies SET notes = ? WHERE id = ?", (comment, vid))
                if should_delete:
                    conn.execute(
                        "UPDATE vacancies SET deleted_at = datetime('now'), delete_reason = ? WHERE id = ?",
                        (del_reason, vid)
                    )
            report['created'] += 1

        # Create contact with priority S when status becomes invited
        if hr_name and ((vh_status == 'invited') or (should_delete and existing and existing['status'] == 'invited')) and vid:
            if not dry_run:
                company_id = _resolve_company_id(conn, company_name) if company_name else None
                if company_id:
                    existing_c = conn.execute(
                        "SELECT id FROM contacts WHERE name = ? AND company_id = ?", (hr_name, company_id)
                    ).fetchone()
                else:
                    existing_c = conn.execute(
                        "SELECT id FROM contacts WHERE name = ? AND company_id IS NULL", (hr_name,)
                    ).fetchone()
                if not existing_c:
                    conn.execute(
                        "INSERT INTO contacts (company_id, name, role, source, priority) VALUES (?, ?, ?, ?, ?)",
                        (company_id, hr_name, 'HR', 'hh.ru', 'S')
                    )
                    report['contacts_created'] += 1

    if not dry_run:
        conn.commit()
    conn.close()
    return report


def main() -> int:
    dry_run = '--dry-run' in sys.argv
    sheet_name = None
    for i, a in enumerate(sys.argv):
        if a == '--sheet' and i + 1 < len(sys.argv):
            sheet_name = sys.argv[i + 1]
    report = import_retro(dry_run, sheet_name)

    print()
    print('=' * 50)
    print(f'Sheet: {report.get("sheet", "?")}')
    if dry_run:
        print('*** DRY RUN — no changes made ***')
    print('=' * 50)
    print(f'Всего строк с данными: {report["total_rows"]}')
    print(f'  Пропущено (нет URL):     {report["skipped_no_url"]}')
    print(f'  Создано новых вакансий:  {report["created"]}')
    print(f'  Обновлено вакансий:      {report["updated"]}')
    print(f'  Пропущено (downgrade):   {report["skipped_downgrade"]}')
    print(f'  Пропущено (terminal):    {report["skipped_terminal"]}')
    print(f'  Создано контактов (S):   {report["contacts_created"]}')
    if report.get('errors'):
        print(f'\nОшибки ({len(report["errors"])}):')
        for e in report['errors'][:10]:
            print(f'  - {e}')
        if len(report['errors']) > 10:
            print(f'  ... и ещё {len(report["errors"]) - 10}')
    print()

    if report.get('error'):
        print(f'ERROR: {report["error"]}')
        return 1
    return 0


if __name__ == '__main__':
    sys.exit(main())
