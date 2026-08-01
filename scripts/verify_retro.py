"""
Верификация импорта ретро: сверяет статусы в Excel и VH.
Запуск: python scripts/verify_retro.py
"""
import sys, os, re
sys.path.insert(0, os.path.join(os.path.dirname(__file__), '..'))

import openpyxl
from src.db import get_connection
from src.retro_import import EXCEL_PATH, STATUS_MAP, excel_status_to_vh, should_update_status, PIPELINE, TERMINAL

conn = get_connection()
wb = openpyxl.load_workbook(str(EXCEL_PATH), data_only=True)
sheet = max(wb.sheetnames, key=lambda s: s.split('.')[::-1])
ws = wb[sheet]

print(f'--- Верификация импорта: лист {sheet} ---')
print()

ok = 0
fail = 0
report = []

for row in ws.iter_rows(min_row=16, max_row=ws.max_row, values_only=True):
    company, url, excel_st, comment, hr_name = (row + (None,)*5)[:5]
    if not company and not url:
        continue

    company = str(company).strip() if company else ''
    url = str(url).strip() if url else ''
    excel_st = str(excel_st).strip() if excel_st else ''

    m = re.search(r'vacancy/(\d+)', url) if url else None
    if not m:
        continue

    hh_id = m.group(1)
    vh_status, should_del, del_reason = excel_status_to_vh(excel_st)

    v = conn.execute("SELECT id, status, deleted_at, delete_reason, notes, hr_contacts FROM vacancies WHERE hh_id=?", (hh_id,)).fetchone()
    if not v:
        report.append(f'  [MISS] {company}: vacancy not found in VH (hh_id={hh_id})')
        fail += 1
        continue

    issues = []

    # Check status / delete (accounting for protection rules)
    if v['deleted_at']:
        # Terminal: skip if Excel says delete, otherwise it's expected protection
        if should_del:
            if del_reason and v['delete_reason'] and del_reason not in str(v['delete_reason']):
                issues.append(f'delete_reason mismatch: expected "{del_reason}", got "{v["delete_reason"]}"')
        # else: terminal protection — OK, skip
    elif should_del:
        issues.append(f'expected DELETED, but deleted_at is NULL')
    elif vh_status:
        cur = v['status'] or 'new'
        if not should_update_status(cur, vh_status):
            pass  # downgrade protection — OK, skip
        elif cur != vh_status:
            issues.append(f'status mismatch: expected "{vh_status}", got "{cur}"')

    # Check notes contain comment
    if comment and comment.strip():
        comment_clean = comment.strip()
        notes = v['notes'] or ''
        if comment_clean not in notes:
            issues.append(f'comment not found in notes')

    if issues:
        report.append(f'  [ISSUE] {company} ({hh_id}): {" | ".join(issues)}')
        fail += 1
    else:
        ok += 1

conn.close()

print(f'Проверено: {ok + fail}')
print(f'  OK:     {ok}')
print(f'  ISSUES: {fail}')
print()
if report:
    print('Детали:')
    for r in report:
        print(r)
else:
    print('Всё совпадает!')
